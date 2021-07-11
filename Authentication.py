from openpyxl import load_workbook, Workbook

wb = load_workbook('accounts.xlsx')
ws = wb.active

userin = 'sign up'

if userin == 'sign up':
    username = ''
    for row in ws.iter_rows(min_row = 'A', max_row = 'B' + 1):
        for cell in row:
            if cell.value == username:
                status = 'already exists'
            else:
                status = 'ok'
    eMail_ID = ''
    for row in ws.iter_rows(min_row = 'B', max_row = 'C'):
        for cell in row:
            if cell.value == eMail_ID:
                status = 'E-Mail already exists'
            else:
                status = 'Ok'
    Password = ''
    Unique_ID = eMail_ID +'-'+ password
    ws.append([username, eMail_ID, Password, Unique_ID])
    wb.save('accounts.xlsx')

if userin == 'sign up':
    eMail_ID = ''
    password = ''
    Unique_ID = eMail +'-'+ password
    for row in ws.iter_rows(min_row=D, max_row=E):
        for cell in row:
            if cell.value == Unique_ID:
                status = 'loading home page...'
            else:
                status = 'either your email or password is incorrect'

